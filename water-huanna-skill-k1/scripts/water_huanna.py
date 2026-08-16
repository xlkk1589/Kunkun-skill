#!/usr/bin/env python3
"""Water-Huanna-Skill-K1 command-line engine.

The engine profiles tabular data, recommends chart styles from the 159-profile
catalog, validates a selected style, and renders a conservative subset of chart
families. Unsupported high-complexity renders return a structured block report.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import re
import statistics
import sys
from collections import Counter
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
CATALOG_PATH = ROOT / "references" / "chart-catalog.json"

ROLE_ALIASES = {
    "sample": ["sample", "sample_id", "样本", "样本编号"],
    "subject": ["subject", "subject_id", "patient", "patient_id", "个体", "患者", "受试者"],
    "group": ["group", "treatment", "class", "condition", "组别", "处理", "类型"],
    "facet": ["facet", "panel", "stratum", "分面", "层级"],
    "category": ["category", "class", "type", "term", "类别", "分类", "项目"],
    "value": ["value", "abundance", "response", "measurement", "数值", "值", "丰度", "响应"],
    "x": ["x", "x_value", "predictor", "自变量"],
    "y": ["y", "y_value", "response", "因变量"],
    "time": ["time", "date", "year", "month", "day", "时间", "日期", "年份"],
    "feature": ["feature", "gene", "protein", "otu", "taxon", "变量", "基因", "蛋白"],
    "effect": ["effect", "estimate", "log2fc", "logfc", "beta", "效应", "估计值"],
    "effect_x": ["effect_x", "logfc_x", "comparison_x", "效应x"],
    "effect_y": ["effect_y", "logfc_y", "comparison_y", "效应y"],
    "estimate": ["estimate", "effect", "beta", "hr", "or", "估计值", "效应量"],
    "lower": ["lower", "ci_low", "lcl", "lower_ci", "置信区间下限", "下限"],
    "upper": ["upper", "ci_high", "ucl", "upper_ci", "置信区间上限", "上限"],
    "error_lower": ["error_lower", "ymin", "lower", "se_low", "误差下限"],
    "error_upper": ["error_upper", "ymax", "upper", "se_high", "误差上限"],
    "p_value": ["p", "pvalue", "p_value", "padj", "fdr", "qvalue", "p值", "校正p值"],
    "adjusted_p": ["padj", "fdr", "qvalue", "adjusted_p", "校正p值"],
    "source": ["source", "from", "来源", "起点"],
    "target": ["target", "to", "目标", "终点"],
    "weight": ["weight", "flow", "count", "strength", "权重", "流量", "强度"],
    "set": ["set", "set_name", "集合"],
    "member": ["member", "item", "element", "成员", "元素"],
    "longitude": ["longitude", "lon", "lng", "经度"],
    "latitude": ["latitude", "lat", "纬度"],
    "crs": ["crs", "epsg", "projection", "坐标系", "投影"],
    "component_a": ["component_a", "a", "组分a"],
    "component_b": ["component_b", "b", "组分b"],
    "component_c": ["component_c", "c", "组分c"],
    "size": ["size", "magnitude", "count", "面积", "大小"],
    "position": ["position", "pos", "bp", "位置"],
    "term": ["term", "variable", "factor", "指标", "变量", "因素"],
    "truth": ["truth", "true_value", "真实值", "真值"],
    "parameter": ["parameter", "term", "参数"],
    "spatial_x": ["spatial_x", "imagecol", "pxl_col", "空间x"],
    "spatial_y": ["spatial_y", "imagerow", "pxl_row", "空间y"],
    "expression": ["expression", "expr", "count", "表达量"],
    "start": ["start", "start_time", "开始", "起始时间"],
    "end": ["end", "end_time", "结束", "终止时间"],
    "event": ["event", "status", "outcome", "事件", "状态"],
    "axis": ["axis", "metric", "indicator", "指标"],
    "series": ["series", "group", "样本", "组别"],
    "row": ["row", "row_id", "行"],
    "column": ["column", "col", "column_id", "列"],
    "component": ["component", "part", "组分"],
    "proportion": ["proportion", "ratio", "percent", "比例", "占比"],
    "label": ["label", "name", "名称", "标签"],
}

ROLE_REASON = {
    "subject": "配对或重复测量必须识别同一观察对象",
    "group": "组间比较需要明确处理或类别",
    "time": "时间序列需要可排序的时间字段",
    "estimate": "森林图必须提供点估计",
    "lower": "不确定性图需要置信区间下限",
    "upper": "不确定性图需要置信区间上限",
    "p_value": "推断图需要显著性或多重校正结果",
    "source": "关系或流图需要起点",
    "target": "关系或流图需要终点",
    "weight": "关系或流图需要边权或流量",
    "crs": "空间图需要明确坐标参考系",
    "newick_tree": "系统发育图需要 Newick 树结构",
    "tip_annotation": "树图附加图层需要与叶节点一一对应的注释",
    "feature_matrix": "降维分析需要样本乘特征矩阵或距离矩阵",
    "numeric_features": "相关矩阵需要多个连续变量",
    "annotation_track": "多轨环图需要逐轨注释",
    "error_lower": "误差线需要下界或可计算误差的重复观测",
    "error_upper": "误差线需要上界或可计算误差的重复观测",
}

GOAL_KEYWORDS = {
    "comparison": ["比较", "差异", "组间", "处理", "comparison", "difference"],
    "distribution": ["分布", "离群", "偏态", "distribution", "outlier"],
    "association": ["相关", "关联", "回归", "association", "correlation", "regression"],
    "time": ["时间", "趋势", "变化", "time", "trend"],
    "space": ["空间", "地图", "地理", "spatial", "map"],
    "effect": ["效应", "置信区间", "meta", "hr", "or", "effect"],
    "composition": ["组成", "占比", "比例", "composition", "proportion"],
    "network": ["网络", "流向", "关系", "network", "flow"],
    "omics": ["基因", "蛋白", "otu", "差异表达", "omics", "gene"],
}

FAMILY_GOALS = {
    "comparison": {"bar", "boxplot", "violin", "raincloud", "forest", "radar"},
    "distribution": {"raincloud", "violin", "boxplot", "ridge", "density_contour"},
    "association": {"scatter", "bubble", "correlation_matrix", "network", "differential_concordance"},
    "time": {"time_series", "line", "swimmer"},
    "space": {"map", "spatial_transcriptomics"},
    "effect": {"forest", "volcano", "parameter_recovery"},
    "composition": {"ternary", "bar", "circular_bar", "scatter_pie", "chord"},
    "network": {"network", "sankey_alluvial", "chord", "upset_venn"},
    "omics": {"volcano", "heatmap", "ordination", "phylogenetic_tree", "manhattan"},
}

COMPLEX_FAMILIES = {
    "map", "network", "sankey_alluvial", "phylogenetic_tree", "chord",
    "spatial_transcriptomics", "circular_sample_track", "scatter_pie",
    "dot_pie_composite", "ordination", "upset_venn", "ternary",
}


def norm_name(value: str) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", value.lower())


def parse_scalar(value: Any) -> Any:
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text or text.lower() in {"na", "nan", "null", "none", "n/a"}:
        return None
    try:
        number = float(text.replace(",", ""))
        return int(number) if number.is_integer() else number
    except ValueError:
        return text


def load_records(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        delimiter = "\t" if suffix in {".tsv", ".txt"} else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [{key: parse_scalar(value) for key, value in row.items()} for row in csv.DictReader(handle, delimiter=delimiter)]
    if suffix in {".json", ".geojson"}:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if suffix == ".geojson" and isinstance(payload, dict) and payload.get("type") == "FeatureCollection":
            rows = []
            for feature in payload.get("features", []):
                row = dict(feature.get("properties") or {})
                geometry = feature.get("geometry") or {}
                coordinates = geometry.get("coordinates")
                if geometry.get("type") == "Point" and isinstance(coordinates, list) and len(coordinates) >= 2:
                    row["longitude"], row["latitude"] = coordinates[:2]
                rows.append(row)
            return rows
        if isinstance(payload, list):
            return [dict(row) for row in payload]
        if isinstance(payload, dict):
            for key in ("records", "data", "rows"):
                if isinstance(payload.get(key), list):
                    return [dict(row) for row in payload[key]]
        raise ValueError("JSON 须为对象数组，或含 records、data、rows 数组")
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("读取 XLSX 需要 openpyxl") from exc
        book = load_workbook(path, read_only=True, data_only=True)
        sheet = book.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else f"column_{index + 1}" for index, value in enumerate(next(rows))]
        return [{key: parse_scalar(value) for key, value in zip(headers, row)} for row in rows]
    raise ValueError(f"暂不支持 {suffix or '无扩展名'}，请转换为 CSV、TSV、JSON、GeoJSON 或 XLSX")


def is_datetime(value: Any) -> bool:
    if isinstance(value, (dt.date, dt.datetime)):
        return True
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not re.search(r"[-/年:]", text):
        return False
    try:
        dt.datetime.fromisoformat(text.replace("年", "-").replace("月", "-").replace("日", ""))
        return True
    except ValueError:
        return False


def profile_records(records: list[dict[str, Any]]) -> dict[str, Any]:
    if not records:
        raise ValueError("数据为空")
    columns = list(dict.fromkeys(key for row in records for key in row))
    details = {}
    numeric_columns, datetime_columns, categorical_columns = [], [], []
    for column in columns:
        values = [row.get(column) for row in records]
        valid = [value for value in values if value is not None]
        numeric = [float(value) for value in valid if isinstance(value, (int, float)) and not isinstance(value, bool)]
        temporal = [value for value in valid if is_datetime(value)]
        unique = len({str(value) for value in valid})
        if valid and len(numeric) / len(valid) >= 0.8:
            inferred = "numeric"
            numeric_columns.append(column)
        elif valid and len(temporal) / len(valid) >= 0.8:
            inferred = "datetime"
            datetime_columns.append(column)
        else:
            inferred = "categorical"
            categorical_columns.append(column)
        details[column] = {
            "type": inferred,
            "missing": len(values) - len(valid),
            "unique": unique,
            "min": min(numeric) if numeric else None,
            "max": max(numeric) if numeric else None,
        }
    role_map = infer_roles(columns, details)
    return {
        "rows": len(records),
        "columns": len(columns),
        "column_order": columns,
        "column_details": details,
        "numeric_columns": numeric_columns,
        "datetime_columns": datetime_columns,
        "categorical_columns": categorical_columns,
        "inferred_roles": role_map,
        "matrix_like": len(numeric_columns) >= 3,
        "has_repeated_subjects": _has_repeats(records, role_map.get("subject")),
    }


def infer_roles(columns: list[str], details: dict[str, dict[str, Any]]) -> dict[str, str]:
    normalized = {column: norm_name(column) for column in columns}
    assigned = {}
    for role, aliases in ROLE_ALIASES.items():
        alias_keys = [norm_name(alias) for alias in aliases]
        exact = [column for column, name in normalized.items() if name in alias_keys]
        partial_aliases = [alias for alias in alias_keys if len(alias) >= 3]
        partial = [column for column, name in normalized.items() if any(name.startswith(alias) or alias in name for alias in partial_aliases)]
        choices = exact or partial
        if choices:
            assigned[role] = choices[0]
    numeric = [column for column in columns if details[column]["type"] == "numeric"]
    categorical = [column for column in columns if details[column]["type"] == "categorical"]
    temporal = [column for column in columns if details[column]["type"] == "datetime"]
    if temporal and "time" not in assigned:
        assigned["time"] = temporal[0]
    if numeric:
        assigned.setdefault("value", numeric[-1])
        assigned.setdefault("y", numeric[-1])
        assigned.setdefault("x", numeric[0])
    if categorical:
        assigned.setdefault("group", categorical[0])
        assigned.setdefault("category", assigned.get("group", categorical[0]))
        assigned.setdefault("sample", categorical[0])
    return assigned


def _has_repeats(records: list[dict[str, Any]], column: str | None) -> bool:
    if not column:
        return False
    values = [row.get(column) for row in records if row.get(column) is not None]
    return len(values) != len(set(map(str, values)))


def load_catalog() -> list[dict[str, Any]]:
    return json.loads(CATALOG_PATH.read_text(encoding="utf-8"))


def parse_mapping(values: list[str] | None) -> dict[str, str]:
    result = {}
    for value in values or []:
        if "=" not in value:
            raise ValueError(f"字段映射须采用 role=column，当前为 {value}")
        role, column = value.split("=", 1)
        result[role.strip()] = column.strip()
    return result


def resolve_role(role: str, profile: dict[str, Any], explicit: dict[str, str]) -> str | None:
    columns = set(profile["column_order"])
    if role in explicit and explicit[role] in columns:
        return explicit[role]
    inferred = profile["inferred_roles"].get(role)
    if inferred in columns:
        return inferred
    if role in {"numeric_features", "feature_matrix"} and profile["matrix_like"]:
        return "<multiple numeric columns>"
    if role == "newick_tree":
        return explicit.get(role)
    if role in {"tip_annotation", "annotation_track"}:
        return explicit.get(role)
    if role == "crs" and "crs" in explicit:
        return explicit["crs"]
    return None


def goal_classes(goal: str) -> set[str]:
    text = goal.lower()
    return {name for name, keywords in GOAL_KEYWORDS.items() if any(keyword.lower() in text for keyword in keywords)}


def evaluate_style(style: dict[str, Any], profile: dict[str, Any], goal: str, explicit: dict[str, str]) -> dict[str, Any]:
    resolved, missing = {}, []
    for role in style["required_roles"]:
        column = resolve_role(role, profile, explicit)
        if column:
            resolved[role] = column
        else:
            missing.append({
                "role": role,
                "reason": ROLE_REASON.get(role, "该图形编码需要此字段或结构"),
                "accepted_examples": ROLE_ALIASES.get(role, [role]),
            })
    coverage = (len(style["required_roles"]) - len(missing)) / max(1, len(style["required_roles"]))
    score = 50 * coverage
    goals = goal_classes(goal)
    if any(style["family"] in FAMILY_GOALS.get(goal_class, set()) for goal_class in goals):
        score += 25
    if style["family"] in {"raincloud", "boxplot", "violin"} and profile["rows"] >= 20:
        score += 8
    if style["family"] in {"correlation_matrix", "heatmap", "ordination"} and profile["matrix_like"]:
        score += 10
    if style["family"] == "time_series" and profile["datetime_columns"]:
        score += 10
    if style["family"] == "map" and {"longitude", "latitude"}.issubset(resolved):
        score += 10
    if style["fidelity_risks"]:
        score -= min(8, len(style["fidelity_risks"]))
    status = "eligible" if not missing else "blocked"
    return {
        "style_id": style["id"],
        "title": style["title"],
        "family": style["family"],
        "score": round(max(0, min(100, score)), 2),
        "status": status,
        "resolved_roles": resolved,
        "missing": missing,
        "fidelity_risks": style["fidelity_risks"],
        "renderer_hint": style["renderer_hint"],
    }


def recommend(path: Path, goal: str, top: int, mapping: dict[str, str]) -> dict[str, Any]:
    records = load_records(path)
    profile = profile_records(records)
    candidates = [evaluate_style(style, profile, goal, mapping) for style in load_catalog()]
    candidates.sort(key=lambda item: (item["status"] == "eligible", item["score"]), reverse=True)
    eligible = [item for item in candidates if item["status"] == "eligible"]
    return {
        "data": str(path),
        "goal": goal,
        "profile": profile,
        "recommendations": eligible[:top],
        "blocked_examples": [item for item in candidates if item["status"] == "blocked"][:top],
        "decision": "ready" if eligible else "blocked",
        "message": "已按科学适配、字段完备和复现风险排序" if eligible else "当前字段无法满足任何风格的硬约束",
    }


def fidelity_requirements(level: str, style: dict[str, Any], spec: dict[str, Any] | None) -> list[dict[str, str]]:
    spec = spec or {}
    missing = []
    if level in {"l2", "l3"}:
        for field, reason in {
            "font_family": "视觉高保真需要原字体或明确替代字体",
            "palette": "视觉高保真需要精确颜色值及映射范围",
            "theme": "视觉高保真需要背景、网格、边框和轴样式",
            "annotation_rules": "视觉高保真需要标签筛选与避让规则",
            "canvas": "视觉高保真需要宽、高、DPI和边距",
        }.items():
            if field not in spec:
                missing.append({"item": field, "reason": reason})
    if level == "l3":
        for field, reason in {
            "original_code": "像素级复现需要原始绘图代码",
            "original_data": "像素级复现需要生成参考图的数据版本",
            "package_versions": "像素级复现需要完整依赖版本",
            "font_files": "像素级复现需要字体文件或可验证字体版本",
            "random_seed": "随机布局、抖动和抽样需要原随机种子",
            "output_device": "不同图形设备会改变字体度量和抗锯齿",
        }.items():
            if field not in spec:
                missing.append({"item": field, "reason": reason})
    if style["family"] == "map":
        for field, reason in {
            "boundary_source": "地图需要合法且可追溯的边界数据",
            "projection": "地图需要明确投影与坐标参考系",
            "compliance_review": "涉及中国地图公开发布时需要地图合规核验",
        }.items():
            if field not in spec:
                missing.append({"item": field, "reason": reason})
    return missing


def validate(path: Path, style_id: int, mapping: dict[str, str], level: str, spec_path: Path | None) -> dict[str, Any]:
    records = load_records(path)
    profile = profile_records(records)
    style = next((item for item in load_catalog() if item["id"] == style_id), None)
    if not style:
        raise ValueError(f"未知 Style ID {style_id}")
    evaluation = evaluate_style(style, profile, "", mapping)
    spec = json.loads(spec_path.read_text(encoding="utf-8")) if spec_path else None
    fidelity_missing = fidelity_requirements(level, style, spec)
    ready = evaluation["status"] == "eligible" and not fidelity_missing
    return {
        "style": {"id": style_id, "title": style["title"], "family": style["family"]},
        "target_fidelity": level,
        "data_validation": evaluation,
        "fidelity_missing": fidelity_missing,
        "decision": "ready" if ready else "blocked",
        "reason": "满足目标复现等级" if ready else "缺少必要字段、统计前提或视觉规格",
    }


def numeric_values(records: list[dict[str, Any]], column: str) -> list[float]:
    return [float(row[column]) for row in records if isinstance(row.get(column), (int, float))]


def render_basic(path: Path, style_id: int, output: Path, mapping: dict[str, str]) -> dict[str, Any]:
    report = validate(path, style_id, mapping, "l1", None)
    if report["decision"] != "ready":
        return {"render_status": "blocked", "validation": report}
    style = next(item for item in load_catalog() if item["id"] == style_id)
    if style["family"] in COMPLEX_FAMILIES:
        return {
            "render_status": "blocked",
            "reason": "该图型依赖专用布局、模型结果或外部结构文件，基础渲染器不会以近似图冒充高保真结果",
            "missing": style["fidelity_risks"],
            "next_action": "依据 style 档案生成专用 R 或 Python 实现，并提供完整视觉规格后执行 L2 验收",
        }
    try:
        import matplotlib.pyplot as plt
        import numpy as np
        from matplotlib import font_manager
    except ImportError as exc:
        return {"render_status": "blocked", "reason": f"基础渲染需要 matplotlib 与 numpy: {exc}"}
    available_fonts = {font.name for font in font_manager.fontManager.ttflist}
    for candidate in ("Microsoft YaHei", "DengXian", "SimHei", "Noto Sans CJK SC", "DejaVu Sans"):
        if candidate in available_fonts:
            plt.rcParams["font.sans-serif"] = [candidate]
            break
    plt.rcParams["axes.unicode_minus"] = False
    records = load_records(path)
    profile = profile_records(records)
    roles = report["data_validation"]["resolved_roles"]
    family = style["family"]
    fig, ax = plt.subplots(figsize=(7.2, 5.2), dpi=160)
    palette = ["#0B7285", "#E67700", "#6741D9", "#2B8A3E", "#C92A2A"]
    if family in {"scatter", "bubble", "differential_concordance", "parameter_recovery"}:
        x_col = roles.get("x") or roles.get("effect_x") or roles.get("truth")
        y_col = roles.get("y") or roles.get("effect_y") or roles.get("estimate")
        xs, ys = [], []
        for row in records:
            if isinstance(row.get(x_col), (int, float)) and isinstance(row.get(y_col), (int, float)):
                xs.append(float(row[x_col])); ys.append(float(row[y_col]))
        sizes = 40
        if family == "bubble" and roles.get("size"):
            raw = numeric_values(records, roles["size"])
            if raw:
                max_value = max(raw) or 1
                sizes = [30 + 170 * value / max_value for value in raw[:len(xs)]]
        ax.scatter(xs, ys, s=sizes, c=palette[0], alpha=0.68, edgecolors="white", linewidths=0.5)
        if family == "parameter_recovery" or "参数恢复" in style["title"]:
            low, high = min(xs + ys), max(xs + ys)
            ax.plot([low, high], [low, high], linestyle="--", color="#495057")
        ax.set_xlabel(x_col); ax.set_ylabel(y_col)
    elif family == "volcano":
        x_col, p_col = roles["effect"], roles["p_value"]
        xs, ys, colors = [], [], []
        for row in records:
            if isinstance(row.get(x_col), (int, float)) and isinstance(row.get(p_col), (int, float)) and row[p_col] > 0:
                x = float(row[x_col]); y = -math.log10(float(row[p_col])); xs.append(x); ys.append(y)
                colors.append("#C92A2A" if x > 1 and row[p_col] < 0.05 else "#1864AB" if x < -1 and row[p_col] < 0.05 else "#ADB5BD")
        ax.scatter(xs, ys, c=colors, alpha=0.7, s=24)
        ax.axvline(-1, color="#868E96", linestyle="--"); ax.axvline(1, color="#868E96", linestyle="--")
        ax.axhline(-math.log10(0.05), color="#868E96", linestyle="--")
        ax.set_xlabel(x_col); ax.set_ylabel("-log10(p)")
    elif family == "forest":
        term, estimate, lower, upper = (roles[key] for key in ("term", "estimate", "lower", "upper"))
        valid = [row for row in records if all(isinstance(row.get(column), (int, float)) for column in (estimate, lower, upper))]
        labels = [str(row.get(term, index + 1)) for index, row in enumerate(valid)]
        y = np.arange(len(valid))
        values = np.array([float(row[estimate]) for row in valid])
        lo = np.array([float(row[lower]) for row in valid]); hi = np.array([float(row[upper]) for row in valid])
        ax.errorbar(values, y, xerr=[values - lo, hi - values], fmt="o", color=palette[0], ecolor="#495057", capsize=3)
        null_value = 1 if re.search(r"\b(?:HR|OR)\b", style["title"], flags=re.I) else 0
        ax.axvline(null_value, color="#868E96", linestyle="--")
        ax.set_yticks(y, labels); ax.invert_yaxis(); ax.set_xlabel(estimate)
    elif family in {"bar", "circular_bar"}:
        category, value = roles["category"], roles["value"]
        grouped = {}
        for row in records:
            if isinstance(row.get(value), (int, float)):
                grouped.setdefault(str(row.get(category)), []).append(float(row[value]))
        labels = list(grouped); values = [statistics.mean(grouped[label]) for label in labels]
        if family == "circular_bar":
            plt.close(fig); fig, ax = plt.subplots(figsize=(6.5, 6.5), dpi=160, subplot_kw={"projection": "polar"})
            theta = np.linspace(0, 2 * np.pi, len(labels), endpoint=False)
            ax.bar(theta, values, width=2 * np.pi / max(1, len(labels)) * 0.78, color=palette[0], alpha=0.82)
            ax.set_xticks(theta, labels)
        else:
            ax.bar(labels, values, color=palette[0], width=0.7)
            ax.tick_params(axis="x", rotation=35)
        ax.set_ylabel(value)
    elif family in {"line", "time_series"}:
        x_col = roles.get("time") or roles.get("x")
        y_col = roles.get("value") or roles.get("y")
        group_col = roles.get("group")
        groups = {}
        for row in records:
            if isinstance(row.get(y_col), (int, float)):
                groups.setdefault(str(row.get(group_col, "all")), []).append((row.get(x_col), float(row[y_col])))
        for index, (group, pairs) in enumerate(groups.items()):
            pairs.sort(key=lambda pair: str(pair[0]))
            ax.plot([pair[0] for pair in pairs], [pair[1] for pair in pairs], marker="o", label=group, color=palette[index % len(palette)])
        if len(groups) > 1:
            ax.legend(frameon=False)
        ax.set_xlabel(x_col); ax.set_ylabel(y_col); ax.tick_params(axis="x", rotation=35)
    elif family in {"boxplot", "violin", "raincloud", "ridge"}:
        group_col, value_col = roles.get("group"), roles.get("value")
        groups = {}
        for row in records:
            if isinstance(row.get(value_col), (int, float)):
                groups.setdefault(str(row.get(group_col)), []).append(float(row[value_col]))
        labels, values = list(groups), list(groups.values())
        if family == "boxplot":
            ax.boxplot(values, tick_labels=labels, patch_artist=True, boxprops={"facecolor": "#74C0FC"})
        else:
            parts = ax.violinplot(values, showmeans=False, showmedians=True)
            for body in parts["bodies"]:
                body.set_facecolor(palette[0]); body.set_alpha(0.55)
            ax.boxplot(values, positions=range(1, len(values) + 1), widths=0.12, showfliers=False)
            ax.set_xticks(range(1, len(labels) + 1), labels)
        ax.set_ylabel(value_col)
    elif family in {"heatmap", "correlation_matrix"}:
        numeric_columns = profile["numeric_columns"]
        matrix = np.array([[float(row[column]) if isinstance(row.get(column), (int, float)) else np.nan for column in numeric_columns] for row in records])
        if family == "correlation_matrix":
            matrix = np.corrcoef(np.nan_to_num(matrix, nan=np.nanmean(matrix, axis=0)), rowvar=False)
            ylabels = xlabels = numeric_columns
        else:
            ylabels = [str(index + 1) for index in range(matrix.shape[0])]
            xlabels = numeric_columns
        image = ax.imshow(matrix, aspect="auto", cmap="RdBu_r")
        ax.set_xticks(range(len(xlabels)), xlabels, rotation=45, ha="right")
        if len(ylabels) <= 30:
            ax.set_yticks(range(len(ylabels)), ylabels)
        fig.colorbar(image, ax=ax, shrink=0.78)
    elif family == "radar":
        axis_col, value_col = roles["axis"], roles["value"]
        series_col = roles.get("sample") or roles.get("series")
        groups = {}
        for row in records:
            if isinstance(row.get(value_col), (int, float)):
                groups.setdefault(str(row.get(series_col)), {})[str(row.get(axis_col))] = float(row[value_col])
        axes = list(dict.fromkeys(str(row.get(axis_col)) for row in records))
        plt.close(fig); fig, ax = plt.subplots(figsize=(6.5, 6.5), dpi=160, subplot_kw={"projection": "polar"})
        theta = np.linspace(0, 2 * np.pi, len(axes), endpoint=False).tolist(); theta += theta[:1]
        for index, (group, mapping_values) in enumerate(groups.items()):
            values = [mapping_values.get(axis, 0) for axis in axes]; values += values[:1]
            ax.plot(theta, values, linewidth=2, label=group, color=palette[index % len(palette)])
            ax.fill(theta, values, alpha=0.12, color=palette[index % len(palette)])
        ax.set_xticks(theta[:-1], axes); ax.legend(frameon=False, loc="upper right", bbox_to_anchor=(1.25, 1.1))
    else:
        plt.close(fig)
        return {"render_status": "blocked", "reason": f"基础渲染器尚未实现 family={family}", "next_action": "按档案生成专用实现"}
    ax.set_title(f"Style {style_id:03d}  {style['title']}", loc="left", fontsize=11, fontweight="bold")
    for spine in ("top", "right"):
        if hasattr(ax, "spines") and spine in ax.spines:
            ax.spines[spine].set_visible(False)
    fig.tight_layout()
    output.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output, bbox_inches="tight")
    plt.close(fig)
    return {"render_status": "completed", "output": str(output), "style_id": style_id, "fidelity": "L1 scientific-equivalence draft"}


def write_result(result: dict[str, Any], output: Path | None) -> None:
    text = json.dumps(result, ensure_ascii=False, indent=2)
    if output:
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(text + "\n", encoding="utf-8")
    print(text)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Water-Huanna-Skill-K1 科研绘图推荐与复现闸门")
    subparsers = parser.add_subparsers(dest="command", required=True)
    inspect_parser = subparsers.add_parser("inspect", help="识别字段类型与语义角色")
    inspect_parser.add_argument("data", type=Path)
    inspect_parser.add_argument("--json-out", type=Path)
    recommend_parser = subparsers.add_parser("recommend", help="从159个风格中推荐候选")
    recommend_parser.add_argument("data", type=Path)
    recommend_parser.add_argument("--goal", default="探索数据并选择科学解释最稳健的图形")
    recommend_parser.add_argument("--top", type=int, default=5)
    recommend_parser.add_argument("--map", action="append", dest="mapping")
    recommend_parser.add_argument("--json-out", type=Path)
    validate_parser = subparsers.add_parser("validate", help="验证指定风格的数据与复现等级")
    validate_parser.add_argument("data", type=Path)
    validate_parser.add_argument("--style-id", type=int, required=True)
    validate_parser.add_argument("--map", action="append", dest="mapping")
    validate_parser.add_argument("--target-fidelity", choices=["l1", "l2", "l3"], default="l1")
    validate_parser.add_argument("--fidelity-spec", type=Path)
    validate_parser.add_argument("--json-out", type=Path)
    render_parser = subparsers.add_parser("render", help="渲染基础图族或返回阻断理由")
    render_parser.add_argument("data", type=Path)
    render_parser.add_argument("--style-id", type=int, required=True)
    render_parser.add_argument("--output", type=Path, required=True)
    render_parser.add_argument("--map", action="append", dest="mapping")
    catalog_parser = subparsers.add_parser("catalog", help="查询样式目录")
    catalog_parser.add_argument("--family")
    catalog_parser.add_argument("--id", type=int, dest="style_id")
    catalog_parser.add_argument("--json-out", type=Path)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        if args.command == "inspect":
            result = profile_records(load_records(args.data))
            write_result(result, args.json_out)
        elif args.command == "recommend":
            result = recommend(args.data, args.goal, args.top, parse_mapping(args.mapping))
            write_result(result, args.json_out)
        elif args.command == "validate":
            result = validate(args.data, args.style_id, parse_mapping(args.mapping), args.target_fidelity, args.fidelity_spec)
            write_result(result, args.json_out)
            return 0 if result["decision"] == "ready" else 2
        elif args.command == "render":
            result = render_basic(args.data, args.style_id, args.output, parse_mapping(args.mapping))
            write_result(result, None)
            return 0 if result["render_status"] == "completed" else 2
        elif args.command == "catalog":
            styles = load_catalog()
            if args.family:
                styles = [style for style in styles if style["family"] == args.family]
            if args.style_id:
                styles = [style for style in styles if style["id"] == args.style_id]
            write_result({"count": len(styles), "styles": styles}, args.json_out)
        return 0
    except (OSError, ValueError, RuntimeError, json.JSONDecodeError) as exc:
        print(json.dumps({"status": "error", "message": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
